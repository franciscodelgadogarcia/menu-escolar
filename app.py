from flask import Flask, jsonify, send_from_directory, request, Response
import os
import csv
from openpyxl import load_workbook
import unicodedata
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
PLATOS = []
BASE_NUTRICIONAL = {}

# ==============================
# CARGAR BASE NUTRICIONAL
# ==============================
def cargar_base_nutricional():
    global BASE_NUTRICIONAL
    ruta = "ingredientes.csv"
    if os.path.exists(ruta):
        try:
            with open(ruta, mode='r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                required_cols = ["Alimento", "E (Kcal)", "L√≠p (g)", "AGS (g)", "Prot (g)", "HdeC (g)", "Azucares", "Vit A (¬µg)", "Vit C (mg)", "vit D (¬µg)", "Ca (mg)", "Fe (mg)", "Sal"]
                if not all(col in reader.fieldnames for col in required_cols):
                    print("‚ùå CSV no tiene columnas requeridas.")
                    return
                for row in reader:
                    nombre = normalizar_texto(row["Alimento"])
                    if nombre:
                        BASE_NUTRICIONAL[nombre] = {
                            "kcal": float(row["E (Kcal)"]),
                            "lip": float(row["L√≠p (g)"]),
                            "ags": float(row["AGS (g)"]),
                            "prot": float(row["Prot (g)"]),
                            "hdec": float(row["HdeC (g)"]),
                            "azucares": float(row["Azucares"]),
                            "vit_a": float(row["Vit A (¬µg)"]),
                            "vit_c": float(row["Vit C (mg)"]),
                            "vit_d": float(row["vit D (¬µg)"]),
                            "ca": float(row["Ca (mg)"]),
                            "fe": float(row["Fe (mg)"]),
                            "sal": float(row["Sal"])
                        }
            print(f"‚úÖ Base nutricional cargada: {len(BASE_NUTRICIONAL)} ingredientes")
        except Exception as e:
            print(f"‚ö†Ô∏è Error al cargar ingredientes.csv: {e}")
    else:
        print("‚ö†Ô∏è ingredientes.csv no encontrado.")

def normalizar_texto(texto):
    if not texto:
        return ""
    texto = texto.lower().strip()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return texto.replace("√°", "a").replace("√©", "e").replace("√≠", "i").replace("√≥", "o").replace("√∫", "u")

def buscar_ingrediente(nombre_ing):
    nombre_norm = normalizar_texto(nombre_ing)
    if not nombre_norm:
        return BASE_NUTRICIONAL.get("agua", {"kcal":0,"lip":0,"ags":0,"prot":0,"hdec":0,"azucares":0,"vit_a":0,"vit_c":0,"vit_d":0,"ca":0,"fe":0,"sal":0})

    # 1. Coincidencia exacta
    if nombre_norm in BASE_NUTRICIONAL:
        return BASE_NUTRICIONAL[nombre_norm]

    # 2. Buscar por palabras clave (m√°s robusto)
    palabras_ing = set(nombre_norm.split())
    mejor_coincidencia = None
    max_coincidencias = 0

    for clave_norm, nut in BASE_NUTRICIONAL.items():
        palabras_clave = set(clave_norm.split())
        coincidencias = len(palabras_ing & palabras_clave)  # intersecci√≥n
        
        # Tambi√©n buscar subcadenas de al menos 3 letras
        for palabra in palabras_ing:
            if len(palabra) >= 3:
                for palabra_clave in palabras_clave:
                    if palabra in palabra_clave or palabra_clave in palabra:
                        coincidencias += 1
        
        if coincidencias > max_coincidencias:
            max_coincidencias = coincidencias
            mejor_coincidencia = nut

    if mejor_coincidencia and max_coincidencias >= 1:
        print(f"üîç Coincidencia por palabras: '{nombre_ing}' ‚Üí '{mejor_coincidencia}' ({max_coincidencias} coincidencias)")
        return mejor_coincidencia

    # 3. √öltimo recurso: usar "agua"
    print(f"‚ö†Ô∏è Ingrediente no encontrado, usando 'agua': '{nombre_ing}'")
    return BASE_NUTRICIONAL.get("agua", {"kcal":0,"lip":0,"ags":0,"prot":0,"hdec":0,"azucares":0,"vit_a":0,"vit_c":0,"vit_d":0,"ca":0,"fe":0,"sal":0})

def calcular_nutricion_plato(ingredientes_gramaje):
    total = {"kcal": 0, "lip": 0, "ags": 0, "prot": 0, "hdec": 0, "azucares": 0, "vit_a": 0, "vit_c": 0, "vit_d": 0, "ca": 0, "fe": 0, "sal": 0}
    for item in ingredientes_gramaje:
        nombre = item["nombre"]
        gramos = item["gramos"]
        if gramos <= 0: continue
        nut = buscar_ingrediente(nombre)
        factor = gramos / 100.0
        for k in total:
            total[k] += nut.get(k, 0) * factor
    for k in total:
        total[k] = round(total[k], 1)
    return total

def leer_ficha_tecnica(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    nombre = str(ws["A7"].value).strip() if ws["A7"].value else ""
    ingredientes = str(ws["A10"].value).strip() if ws["A10"].value else ""

    alergenos = []
    alergenos_col1 = ["Gluten", "Crust√°ceos", "Huevos", "Pescado", "Cacahuetes", "Soja", "Leche", "Legumbres", "Guisantes"]
    for i, nombre_alerg in enumerate(alergenos_col1, start=12):
        if ws[f"F{i}"].value == "X":
            alergenos.append(nombre_alerg)

    alergenos_col2 = ["Frutos de c√°scara", "Apio", "Mostaza", "S√©samo", "Sulfuroso", "Altramuces", "Moluscos", "Cerdo", "Otros"]
    for i, nombre_alerg in enumerate(alergenos_col2, start=12):
        if ws[f"K{i}"].value == "X":
            alergenos.append(nombre_alerg)

    ingredientes_gramaje = []
    for fila in range(35, 44):
        nombre_ing = ws[f"E{fila}"].value
        gramos = ws[f"H{fila}"].value
        if nombre_ing and gramos is not None:
            try:
                gramos = float(gramos)
                ingredientes_gramaje.append({"nombre": str(nombre_ing).strip(), "gramos": gramos})
            except (ValueError, TypeError):
                pass

    gramos_racion = float(ws["H44"].value) if isinstance(ws["H44"].value, (int, float)) else sum(item["gramos"] for item in ingredientes_gramaje)
    nutricion = calcular_nutricion_plato(ingredientes_gramaje)

    return {
        "nombre": nombre,
        "ingredientes": ingredientes,
        "alergenos": alergenos,
        "ingredientes_gramaje": ingredientes_gramaje,
        "gramos_racion": gramos_racion,
        "nutricion": nutricion,
        "archivo": os.path.basename(ruta_excel)
    }

def cargar_platos():
    global PLATOS
    PLATOS = []
    for filename in os.listdir("."):
        if filename.endswith(".xlsx") and filename != "plantilla_menu.xlsx":
            try:
                plato = leer_ficha_tecnica(filename)
                PLATOS.append(plato)
                print(f"‚úÖ Cargado: {filename}")
            except Exception as e:
                print(f"‚ùå Error al cargar {filename}: {e}")

# ==============================
# RUTA PARA EXPORTAR USANDO PLANTILLA
# ==============================
@app.route("/api/exportar", methods=["POST"])
def exportar_menu():
    data = request.json
    colegio = data.get("colegio")
    mes = data.get("mes")
    anio = data.get("anio")
    menu_datos = data.get("menu", {})

    if not colegio or not menu_datos:
        return jsonify({"error": "Faltan datos"}), 400

    # Obtener listas de platos
platos_dict = clasificar_platos_dict()
primeros = platos_dict["primeros"]
segundos = platos_dict["segundos"]
acompanamientos = platos_dict["acompanamientos"]
postres = platos_dict["postres"]
panes = platos_dict["panes"]
todos_platos = primeros + segundos + acompanamientos + postres + panes

    def get_plato(nombre):
        for p in todos_platos:
            if p["nombre"] == nombre:
                return p
        return None

    # Iniciar HTML con soporte para Excel
    html = '''<!DOCTYPE html>
<html xmlns:o="urn:schemas-microsoft-com:office:office" 
      xmlns:x="urn:schemas-microsoft-com:office:excel" 
      xmlns="http://www.w3.org/TR/REC-html40">
<head>
<meta http-equiv=Content-Type content="text/html; charset=utf-8">
<style>
body { font-family: Arial; }
table { border-collapse: collapse; width: 100%; }
td, th { 
    border: 1px solid #000; 
    padding: 4px; 
    font-size: 11pt;
    height: 20px;
}
.encabezado { 
    font-family: "Very Simple Chalk"; 
    font-size: 22pt; 
    color: #385724; 
    text-align: center;
    height: 30px;
}
.celda-combinada {
    text-align: center;
    vertical-align: middle;
    font-size: 16pt;
    height: 25px;
}
.cena { 
    background-color: #ffcc99; 
}
.rotado { 
    writing-mode: tb-rl; 
    text-align: center; 
    vertical-align: middle;
    height: 100px;
    width: 20px;
}
.columna-dia { 
    width: 20px; 
    text-align: center; 
    vertical-align: middle;
    font-size: 12pt;
}
</style>
<!--[if gte mso 9]><xml>
<x:ExcelWorkbook><x:ExcelWorksheets><x:ExcelWorksheet>
<x:Name>Men√∫ Mensual</x:Name>
<x:WorksheetOptions><x:DisplayGridlines/></x:WorksheetOptions>
</x:ExcelWorksheet></x:ExcelWorksheets></x:ExcelWorkbook></xml><![endif]-->
</head>
<body>
<table>
<tr>
    <td colspan="8"></td>
    <td colspan="28" class="encabezado">Men√∫ Escolar - ''' + colegio + '''</td>
</tr>
<tr>
    <td colspan="8"></td>
    <td colspan="20" class="encabezado">''' + mes + '''</td>
    <td colspan="8" class="encabezado">''' + str(anio) + '''</td>
</tr>
'''

    # Generar filas por d√≠a
    dias_procesados = 0
    for fecha_str, menu in menu_datos.items():
        if not any([menu.get("primer"), menu.get("segundo"), menu.get("postre")]):
            continue

        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
        except:
            continue

        dia_semana = fecha.strftime("%a")[:3]
        if dia_semana not in ['Lun', 'Mar', 'Mi√©', 'Jue', 'Vie']:
            continue

        # Obtener platos
        p1 = get_plato(menu.get("primer", ""))
        p2 = get_plato(menu.get("segundo", ""))
        pA = get_plato(menu.get("acompanamiento", ""))
        p3 = get_plato(menu.get("postre", ""))
        pPan = get_plato(menu.get("pan", ""))

        # Calcular nutrici√≥n
        def get_nut(plato):
            return plato["nutricion"] if plato else {"kcal":0,"lip":0,"ags":0,"prot":0,"hdec":0,"azucares":0,"vit_a":0,"vit_c":0,"vit_d":0,"ca":0,"fe":0,"sal":0}

        nut1, nut2, nutA, nut3, nutPan = map(get_nut, [p1, p2, pA, p3, pPan])

        def sumar_nut(clave):
            return round(nut1[clave] + nut2[clave] + nutA[clave] + nut3[clave] + nutPan[clave], 1)

        # Construir filas del d√≠a
        html += f'''
        <tr>
            <td class="columna-dia" rowspan="13">{fecha.day}</td>
            <td class="celda-combinada" colspan="6">{p1["nombre"] if p1 else ""}</td>
        </tr>
        <tr><td class="celda-combinada" colspan="6">{traducir_al_ingles(p1["nombre"] if p1 else "")}</td></tr>
        <tr><td class="celda-combinada" colspan="6">{p2["nombre"] if p2 else ""}</td></tr>
        <tr><td class="celda-combinada" colspan="6">{pA["nombre"] if pA else ""}</td></tr>
        <tr><td class="celda-combinada" colspan="6">{traducir_al_ingles(p2["nombre"] if p2 else "")} with {traducir_al_ingles(pA["nombre"] if pA else "")}</td></tr>
        <tr><td class="celda-combinada" colspan="6">{(p3["nombre"] if p3 else "") + " + Agua + " + (pPan["nombre"] if pPan else "Pan")}</td></tr>
        <tr><td class="celda-combinada" colspan="6">{traducir_al_ingles(p3["nombre"] if p3 else "")} + Water + {traducir_al_ingles(pPan["nombre"] if pPan else "Bread")}</td></tr>
        <tr>
            <td>E (Kcal)</td><td>L√≠p (g)</td><td>AGS (g)</td><td>Prot (g)</td><td>HdeC (g)</td><td>Azucares</td>
        </tr>
        <tr>
            <td>{sumar_nut("kcal")}</td><td>{sumar_nut("lip")}</td><td>{sumar_nut("ags")}</td><td>{sumar_nut("prot")}</td><td>{sumar_nut("hdec")}</td><td>{sumar_nut("azucares")}</td>
        </tr>
        <tr>
            <td>Vit A (¬µg)</td><td>Vit C (mg)</td><td>vit D (¬µg)</td><td>Ca (mg)</td><td>Fe (mg)</td><td>Sal</td>
        </tr>
        <tr>
            <td>{sumar_nut("vit_a")}</td><td>{sumar_nut("vit_c")}</td><td>{sumar_nut("vit_d")}</td><td>{sumar_nut("ca")}</td><td>{sumar_nut("fe")}</td><td>{sumar_nut("sal")}</td>
        </tr>
        <tr>
            <td class="cena">cena</td>
            <td class="cena" colspan="5"></td>
        </tr>
        <tr>
            <td class="cena"></td>
            <td class="cena" colspan="5"></td>
        </tr>
        '''

        dias_procesados += 1
        if dias_procesados >= 25:  # Evitar bucles infinitos
            break

    # Pie de p√°gina
    html += '''
    </table>
    <br><br>
    <div style="font-size: 12pt; font-family: Arial;">
    Valorado nutricionalmente por Leticia Montoiro Peinado, Diplomada en Nutrici√≥n Humana y Diet√©tica. N¬∫ Colegiada CYL00207<br>
    La fruta podr√° variar en funci√≥n de su grado de madurez. Valoraci√≥n nutricional en base a ni√±os de 6-9 a√±os.<br>
    Para cualquier consulta del men√∫ o informaci√≥n de al√©rgenos, puedes enviar un correo a nuestra nutricionista a: nutricion@cofuri.es<br>
    * Las recetas elaboradas llevan excluidos los alimentos arriba detallados.
    </div>
    </body>
    </html>
    '''

    return Response(
        html,
        mimetype="application/vnd.ms-excel",
        headers={"Content-Disposition": f"attachment;filename=menu_escolar_{colegio.replace(' ', '_')}_{mes}_{anio}.xls"}
    )

def clasificar_platos_dict():
    """Versi√≥n de clasificar_platos para uso en backend"""
    primeros = [p for p in PLATOS if p["archivo"].startswith("PR.")]
    segundos = [p for p in PLATOS if p["archivo"].startswith("PO.")]
    acompanamientos = [p for p in PLATOS if p["archivo"].startswith("AC.")]
    postres = [p for p in PLATOS if p["archivo"].startswith("DE.")]
    panes = [p for p in PLATOS if p["archivo"].startswith("PA.")]
    return {
        "primeros": primeros,
        "segundos": segundos,
        "acompanamientos": acompanamientos,
        "postres": postres,
        "panes": panes
    }
    def traducir_al_ingles(texto):
    if not texto:
        return ""
    traducciones = {
        "Lentejas": "Lentils", "Alubias": "Beans", "Sopa": "Soup", "Crema": "Cream", "Pasta": "Pasta",
        "Guiso": "Stew", "Pur√©": "Mash", "Verduras": "Vegetables", "Pollo": "Chicken", "Pescado": "Fish",
        "Carne": "Meat", "Jam√≥n": "Ham", "Chorizo": "Chorizo", "Morcilla": "Blood sausage", "Merluza": "Hake",
        "Bacalao": "Cod", "Salm√≥n": "Salmon", "At√∫n": "Tuna", "Sardinas": "Sardines", "Patata": "Potato",
        "Zanahoria": "Carrot", "Cebolla": "Onion", "Tomate": "Tomato", "Pimiento": "Pepper", "Calabac√≠n": "Zucchini",
        "Espinacas": "Spinach", "Acelgas": "Chard", "Jud√≠as verdes": "Green beans", "Br√≥coli": "Broccoli",
        "Coliflor": "Cauliflower", "Lechuga": "Lettuce", "Puerro": "Leek", "Apio": "Celery", "Remolacha": "Beetroot",
        "Champi√±√≥n": "Mushroom", "Ajo": "Garlic", "Guisantes": "Peas", "Ma√≠z": "Corn", "Manzana": "Apple",
        "Pera": "Pear", "Pl√°tano": "Banana", "Naranja": "Orange", "Mandarina": "Tangerine", "Uva": "Grape",
        "Mel√≥n": "Melon", "Sand√≠a": "Watermelon", "Fresa": "Strawberry", "Kiwi": "Kiwi", "Pi√±a": "Pineapple",
        "Melocot√≥n": "Peach", "Ciruela": "Plum", "Higo": "Fig", "Aguacate": "Avocado", "Leche": "Milk",
        "Yogur": "Yogurt", "Queso fresco": "Fresh cheese", "Reques√≥n": "Cottage cheese", "Huevo": "Egg",
        "Gelatina": "Gelatin", "Flan": "Flan", "Natillas": "Custard", "Pan": "Bread", "Agua": "Water"
    }
    for es, en in traducciones.items():
        if es in texto:
            texto = texto.replace(es, en)
    return texto
# ==============================
# INICIAR APP
# ==============================
cargar_base_nutricional()
cargar_platos()

@app.route("/")
def home():
    return send_from_directory(".", "index.html")

@app.route("/api/platos")
def obtener_platos():
    return jsonify({"platos": PLATOS})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
